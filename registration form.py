import tkinter as tk
from tkinter import ttk
import openpyxl

class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Bano Qabil 3.0 Registration Form")
        self.create_widgets()
        self.set_default_theme()

    def create_widgets(self):
        self.frame = ttk.Frame(self.root)
        self.frame.pack(padx=50, pady=50)

        self.create_insert_widgets()
        
        self.create_treeview()
    

    
    def create_insert_widgets(self):
        widgets_frame = ttk.LabelFrame(self.frame, text="Student Personal Information", width=150)
        widgets_frame.grid(row=0, column=0, padx=25, pady=5, sticky="nsew")
        widgets_frame.columnconfigure(0, weight=1)

        self.name_entry = ttk.Entry(widgets_frame)
        self.name_entry.insert(0, "Type Your Name Here")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.age_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=60)
        self.age_spinbox.insert(0, "Select Your Age")
        self.age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        combo_list = ["Male", "Female"]
        self.gender_combobox = ttk.Combobox(widgets_frame, values=combo_list)
        self.gender_combobox.current(0)
        self.gender_combobox.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")

        self.mobileNo_entry = ttk.Entry(widgets_frame)
        self.mobileNo_entry.insert(0, "Type Your Mobile Number Here")
        self.mobileNo_entry.bind("<FocusIn>", lambda e: self.mobileNo_entry.delete('0', 'end'))
        self.mobileNo_entry.grid(row=3, column=0, padx=5, pady=(0, 10), sticky="ew")

        self.employed_var = tk.BooleanVar()
        self.checkbutton = ttk.Checkbutton(widgets_frame, text="Employed", variable=self.employed_var)
        self.checkbutton.grid(row=4, column=0, padx=5, pady=10, sticky="nsew")

        combo_list = ["Select Course", "The Art Of Digital Marketing & Freelancing", "Mastering Wholesale & Online Arbitrage", "Amazon FBA Private Label Mastery", "Content Writing And ChatGPT", "Sales And Lead Generation", "Thriving With Shopify, WordPress, And Wix", "Cyber Security Fundamentals", "Essentials Of Computing And Programming", "Adobe Illustrator, Photoshop, InDesign, Figma, And Canva", "Video Editing And Animation With Adobe Premiere Pro And After Effects", "Building With HTML, CSS, And JavaScript", "UI/UX Design And Frontend Development With React.JS"]
        self.course_combobox = ttk.Combobox(widgets_frame, values=combo_list, width=30)
        self.course_combobox.current(0)
        self.course_combobox.grid(row=5, column=0, padx=10, pady=10,  sticky="ew")

        self.button = ttk.Button(widgets_frame, text="Insert", command=self.insert_row)
        self.button.grid(row=6, column=0, padx=5, pady=20, sticky="nsew")

        # Make combo_list an instance variable
        self.combo_list = combo_list

    def create_treeview(self):
        treeFrame = ttk.LabelFrame(self.frame, text= "Student's DataBase", width=150)
        treeFrame.grid(row=0, column=1, pady=10)
        treeFrame.columnconfigure(0, weight=1) 
        treeScroll = ttk.Scrollbar(treeFrame)
        treeScroll.pack(side="right", fill="y")

        cols = ("Name", "Age", "Gender", "MobileNo", "Employment","CourseName","Selection Status")
        self.treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=20)
        for col in cols:
            self.treeview.heading(col, text=col, anchor="w")  # east = right, west = left, north = top, south = bottom


        
        self.treeview.column("Name", width=100)
        self.treeview.column("Age", width=50)
        self.treeview.column("Gender", width=100)
        self.treeview.column("MobileNo", width=100)
        self.treeview.column("Employment", width=100)
        self.treeview.column("CourseName", width=200)
        self.treeview.column("Selection Status", width=100)
        self.treeview.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        treeScroll.config(command=self.treeview.yview)

        self.load_data()

    def load_data(self):
        path = "people.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.treeview.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeview.insert('', tk.END, values=value_tuple)
        

    def insert_row(self):
    # def insert_row(self, sheet):
    #     last_serial = 0
    #     for row in sheet.iter_rows(values_only = True):
    #         if row [0] is not None and isinstance (row[0], int):
    #             last_serial = max (last_serial, row[0])
        
        
        # new_serial = last_serial + 1
        name = self.name_entry.get()
        age = int(self.age_spinbox.get())
        gender_status = self.gender_combobox.get()
        MobileNo = self.mobileNo_entry.get()
        employment_status = "Employed" if self.employed_var.get() else "Unemployed"
        CourseName = self.course_combobox.get()

        

        # Update Excel sheet by adding another row at the end
        path = "people.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
    
        
              
        row_values = [name, age, gender_status, MobileNo, employment_status, CourseName]
        sheet.append(row_values)
        workbook.save(path)

        # Insert row into treeview
        self.treeview.insert('', tk.END, values=row_values)

        # Clear the values
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, "Age")
        self.gender_combobox.set(self.combo_list[0])
        self.mobileNo_entry.delete(0, "end")
        self.mobileNo_entry.insert(0, "MobileNo")
        self.employed_var.set(False)
        self.course_combobox.set(self.combo_list[0])

    def set_default_theme(self):
        self.style = ttk.Style(self.root)
        self.root.tk.call("source", "forest-light.tcl")
        self.root.tk.call("source", "forest-dark.tcl")
        self.style.theme_use("forest-dark")


def main():
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()